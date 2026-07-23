"""Carga la produccion historica mensual desde el Excel de data."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
ARCHIVO_PRODUCCION = (
    BASE_DIR / "data" / "P535 - PRODUCCION (CONJUNTO DE PLATAFORMAS).xlsx"
)
HOJA_PRODUCCION = "PRODUCCION"
MAX_FILAS_ENCABEZADO = 15

COLUMNAS_PRODUCCION = {
    "campo",
    "plataforma",
    "pozo",
    "fecha",
    "gas_mmpcd",
    "agua_mbd",
    "aceite_mbd",
    "gor_pc_bbl",
}


def normalizar_texto(valor: Any) -> str:
    return str(valor or "").strip()


def normalizar_nombre_hoja(valor: Any) -> str:
    texto = normalizar_texto(valor).upper()
    for caracter, reemplazo in {
        "Á": "A",
        "É": "E",
        "Í": "I",
        "Ó": "O",
        "Ú": "U",
        "Ñ": "N",
    }.items():
        texto = texto.replace(caracter, reemplazo)
    return " ".join(texto.split())


def normalizar_columna(valor: Any) -> str:
    texto = normalizar_texto(valor).replace("\n", " ").lower()
    texto = " ".join(texto.split())
    equivalencias = {
        "campo": "campo",
        "plataforma": "plataforma",
        "pozo": "pozo",
        "gas mmpcd": "gas_mmpcd",
        "agua mbd": "agua_mbd",
        "aceite mbd": "aceite_mbd",
        "goc pc/bbl": "gor_pc_bbl",
        "gor pc/bbl": "gor_pc_bbl",
        "fecha": "fecha",
    }
    return equivalencias.get(texto, texto)


def normalizar_fecha(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        return date(valor.year, valor.month, 1)
    if isinstance(valor, date):
        return date(valor.year, valor.month, 1)
    if not valor:
        return None

    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            fecha = datetime.strptime(str(valor).strip(), formato)
            return date(fecha.year, fecha.month, 1)
        except ValueError:
            continue
    return None


def normalizar_numero(valor: Any) -> float:
    if valor in (None, ""):
        return 0.0
    try:
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


def encontrar_archivos_produccion(ruta_excel: Path | None = None) -> list[Path]:
    if ruta_excel is not None:
        if ruta_excel.is_dir():
            return sorted(ruta_excel.glob("P535 - PRODUCCION*.xlsx"))
        return [ruta_excel] if ruta_excel.exists() else []

    ruta_base = BASE_DIR / "data"
    if not ruta_base.exists():
        return []

    archivos = sorted(ruta_base.glob("P535 - PRODUCCION*.xlsx"))
    if archivos:
        return archivos

    return [ARCHIVO_PRODUCCION] if ARCHIVO_PRODUCCION.exists() else []


def encontrar_hoja_produccion(libro: Any) -> Any:
    for nombre_hoja in libro.sheetnames:
        nombre_normalizado = normalizar_nombre_hoja(nombre_hoja)
        if HOJA_PRODUCCION in nombre_normalizado:
            return libro[nombre_hoja]
    return libro[libro.sheetnames[0]]


def encontrar_encabezado(hoja: Any) -> tuple[int, dict[str, int]]:
    for fila_numero in range(1, MAX_FILAS_ENCABEZADO + 1):
        encabezados = [
            normalizar_columna(celda.value)
            for celda in hoja[fila_numero]
        ]
        posiciones = {
            nombre: indice
            for indice, nombre in enumerate(encabezados)
            if nombre in COLUMNAS_PRODUCCION
        }
        if COLUMNAS_PRODUCCION.issubset(posiciones):
            return fila_numero, posiciones

    raise RuntimeError(
        "El Excel de produccion no tiene columnas requeridas: "
        + ", ".join(sorted(COLUMNAS_PRODUCCION))
    )


def leer_produccion_excel(
    ruta_excel: Path | None = None,
) -> list[dict[str, Any]]:
    """Lee los Excel de produccion y devuelve registros normalizados."""
    rutas = encontrar_archivos_produccion(ruta_excel)
    if not rutas:
        return []

    registros: dict[tuple[str, str, str, date], dict[str, Any]] = {}

    for ruta in rutas:
        libro = load_workbook(ruta, read_only=True, data_only=True)
        hoja = encontrar_hoja_produccion(libro)
        fila_encabezado, posiciones = encontrar_encabezado(hoja)

        for fila in hoja.iter_rows(min_row=fila_encabezado + 1, values_only=True):
            campo = normalizar_texto(fila[posiciones["campo"]]).upper()
            plataforma = normalizar_texto(fila[posiciones["plataforma"]]).upper()
            pozo = normalizar_texto(fila[posiciones["pozo"]]).upper()
            fecha = normalizar_fecha(fila[posiciones["fecha"]])

            if not campo or not plataforma or not pozo or fecha is None:
                continue

            llave = (campo, plataforma, pozo, fecha)
            registro = {
                "campo": campo,
                "plataforma": plataforma,
                "pozo": pozo,
                "fecha": fecha,
                "gas_mmpcd": normalizar_numero(fila[posiciones["gas_mmpcd"]]),
                "agua_mbd": normalizar_numero(fila[posiciones["agua_mbd"]]),
                "aceite_mbd": normalizar_numero(fila[posiciones["aceite_mbd"]]),
                "gor_pc_bbl": normalizar_numero(fila[posiciones["gor_pc_bbl"]]),
            }

            if llave not in registros:
                registros[llave] = registro
                continue

            registros[llave]["gas_mmpcd"] += registro["gas_mmpcd"]
            registros[llave]["agua_mbd"] += registro["agua_mbd"]
            registros[llave]["aceite_mbd"] += registro["aceite_mbd"]
            registros[llave]["gor_pc_bbl"] = registro["gor_pc_bbl"]

    return sorted(
        registros.values(),
        key=lambda item: (
            item["campo"],
            item["plataforma"],
            item["pozo"],
            item["fecha"],
        ),
    )
